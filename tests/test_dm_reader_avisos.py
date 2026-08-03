"""Lo que antes se ignoraba en silencio al leer el Excel ahora avisa."""
import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from src.excel.dm_reader import leer_dm

AMARILLO_OK = PatternFill("solid", fgColor="FFFF00")
AMARILLO_CASI = PatternFill("solid", fgColor="FFE599")   # "amarillo claro" del tema
GRIS = PatternFill("solid", fgColor="D9D9D9")

HEADERS = ["Documento", "Proveedor", "CUIT", "Comprobante", "Importe", "Forma de pago"]
FILA = ["FC - 21562", "PROVEEDOR SA", "30-71018343-7", "A-0001-00000123", -1000, "Ch 10/09"]


def _libro(filas: list[tuple[list, PatternFill | None]], headers=HEADERS, hoja="DM"):
    """(valores, relleno) por fila → Workbook listo para guardar."""
    wb = Workbook()
    ws = wb.active
    ws.title = hoja
    ws.append(headers)
    for valores, fill in filas:
        ws.append(valores)
        if fill is not None:
            for cell in ws[ws.max_row]:
                cell.fill = fill
    return wb


def _guardar(tmp_path, wb):
    path = tmp_path / "dm.xlsx"
    wb.save(path)
    return path


class TestFilasCasiAmarillas:
    def test_avisa_de_la_fila_con_amarillo_no_estandar(self, tmp_path):
        path = _guardar(tmp_path, _libro([(FILA, AMARILLO_CASI)]))
        avisos: list[str] = []

        proveedores = leer_dm(path, avisos_out=avisos)

        assert proveedores == []          # sigue sin procesarse
        assert len(avisos) == 1
        assert "amarillo estándar" in avisos[0]
        assert "fila 2" in avisos[0]

    def test_amarillo_estandar_se_procesa_sin_avisos(self, tmp_path):
        path = _guardar(tmp_path, _libro([(FILA, AMARILLO_OK)]))
        avisos: list[str] = []

        proveedores = leer_dm(path, avisos_out=avisos)

        assert len(proveedores) == 1
        assert avisos == []

    def test_no_avisa_por_colores_que_no_pretenden_ser_amarillo(self, tmp_path):
        path = _guardar(tmp_path, _libro([(FILA, GRIS)]))
        avisos: list[str] = []

        assert leer_dm(path, avisos_out=avisos) == []
        assert avisos == []

    def test_no_avisa_por_filas_sin_datos_completos(self, tmp_path):
        incompleta = ["FC - 1", "", "", "", -100, "Ch 10/09"]
        path = _guardar(tmp_path, _libro([(incompleta, AMARILLO_CASI)]))
        avisos: list[str] = []

        assert leer_dm(path, avisos_out=avisos) == []
        assert avisos == []

    def test_sin_avisos_out_no_falla(self, tmp_path):
        path = _guardar(tmp_path, _libro([(FILA, AMARILLO_CASI)]))
        assert leer_dm(path) == []


class TestDocumentoNormalizado:
    def test_guion_sin_espacios_queda_canonico(self, tmp_path):
        fila = ["FC-21562", *FILA[1:]]
        path = _guardar(tmp_path, _libro([(fila, AMARILLO_OK)]))

        proveedores = leer_dm(path)

        assert proveedores[0].items[0].documento == "FC - 21562"


class TestColumnas:
    def test_error_de_columnas_indica_la_fila_1_y_lo_leido(self, tmp_path):
        path = _guardar(tmp_path, _libro([(FILA, AMARILLO_OK)],
                                         headers=["Doc", "Prov", "Monto", "Cond pago"]))

        with pytest.raises(ValueError) as exc:
            leer_dm(path)

        mensaje = str(exc.value)
        assert "PRIMERA fila" in mensaje
        assert "Doc, Prov, Monto, Cond pago" in mensaje

    def test_avisa_si_hay_dos_columnas_de_importe(self, tmp_path):
        headers = ["Documento", "Proveedor", "Importe original", "Importe ppal",
                   "Forma de pago"]
        fila = ["FC - 21562", "PROVEEDOR SA", -999, -1000, "Ch 10/09"]
        path = _guardar(tmp_path, _libro([(fila, AMARILLO_OK)], headers=headers))
        avisos: list[str] = []

        proveedores = leer_dm(path, avisos_out=avisos)

        assert any("Importe original" in a and "Se usó" in a for a in avisos)
        # Se usa la de más a la izquierda: -999 en el Excel → +999 interno
        assert proveedores[0].items[0].importe == 999
