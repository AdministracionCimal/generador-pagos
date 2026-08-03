import unicodedata
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from src.domain.clasificador import clasificar
from src.domain.documento import normalizar as normalizar_doc
from src.domain.models import ItemFactura, ProveedorTanda

HOJA_DM = "DM"

# Convención de signos:
#   Excel/Finnegans: negativo = adeuda al proveedor, positivo = saldo a favor
#   Interno (app):   positivo = a pagar (genera cheque/transfer), negativo = crédito
# Por eso invertimos el signo al leer.
#
# Qué filas se procesan: las pintadas de amarillo (solo_amarillas=True).
# El prefijo del documento no decide validez — lo decide la marca del usuario.

def _normalizar(s: str) -> str:
    """Minúsculas y sin tildes para comparar headers."""
    return unicodedata.normalize("NFD", s.lower()).encode("ascii", "ignore").decode()


# Campos que se detectan por coincidencia exacta del header normalizado.
# Evita falsos positivos: "Condicionpago" no debe matchear "pago";
# "Fecha comprobante" no debe matchear "comprobante".
_EXACT = {
    "cuit":          "cuit",
    "documento":     "documento",
    "proveedor":     "proveedor",
    "comprobante":   "comprobante",
    "forma de pago": "pago",   # header preferido
    "pago":          "pago",   # fallback para archivos con header «PAGO»
}

# Campos detectados por contener la clave en cualquier parte del header.
_CONTAINS = {
    "importe":  "importe",
    "fecha vto": "fecha_vto",
}


def _detectar_columnas(header_row, avisos_out: list[str] | None = None) -> dict[str, int]:
    """Devuelve {clave: índice_0based} según los headers de la primera fila.

    Los campos de `_CONTAINS` se resuelven por coincidencia parcial, así que
    puede haber más de un candidato (ej. «Importe original» e «Importe ppal»).
    Se usa el de más a la izquierda y se avisa, porque antes elegía en silencio.
    """
    cols: dict[str, int] = {}
    partial: dict[str, int] = {}
    candidatos: dict[str, list[str]] = {}
    for idx, cell in enumerate(header_row):
        crudo = str(cell.value or "").strip()
        norm = _normalizar(crudo)
        for keyword, key in _EXACT.items():
            if norm == keyword and key not in cols:
                cols[key] = idx
        for keyword, key in _CONTAINS.items():
            if keyword in norm:
                candidatos.setdefault(keyword, []).append(crudo)
                if key not in partial:
                    partial[key] = idx
    for key, idx in partial.items():
        if key not in cols:
            cols[key] = idx

    if avisos_out is not None:
        for keyword, headers in candidatos.items():
            if len(headers) > 1:
                avisos_out.append(
                    f"• Hay {len(headers)} columnas cuyo encabezado contiene "
                    f"«{keyword}» ({', '.join(headers)}). Se usó «{headers[0]}», "
                    f"la primera de izquierda a derecha."
                )
    return cols


def _cel(row, idx: int | None):
    if idx is None or idx >= len(row):
        return None
    return row[idx].value


def _importe(raw) -> Decimal | None:
    if raw is None:
        return None
    try:
        return Decimal(str(raw))
    except InvalidOperation:
        return None


def _fecha(raw) -> date | None:
    if raw is None:
        return None
    if isinstance(raw, date):
        return raw
    return None


def _limpiar_cuit(raw: str) -> str:
    """'30-71018343-7' → '30710183437'"""
    return raw.replace("-", "").replace(".", "").strip()


# Amarillo estándar de Excel, en las codificaciones con y sin canal alfa.
_AMARILLOS_RGB = {"FFFFFF00", "FFFF00", "00FFFF00"}
# Índices del amarillo en la paleta legacy (archivos viejos guardan indexed).
_AMARILLOS_INDEXED = {5, 13}


def _fill_solido(cell):
    """Devuelve el relleno si la celda está efectivamente pintada, o None."""
    fill = getattr(cell, "fill", None)
    if fill is None or not fill.patternType:
        return None
    return fill


def _es_celda_amarilla(cell) -> bool:
    fill = _fill_solido(cell)
    if fill is None:
        return False
    color = fill.fgColor
    if color.type == "rgb" and str(color.rgb or "").upper() in _AMARILLOS_RGB:
        return True
    return color.type == "indexed" and color.indexed in _AMARILLOS_INDEXED


def _es_fila_amarilla(row) -> bool:
    return any(_es_celda_amarilla(c) for c in row)


def _color_casi_amarillo(cell) -> bool:
    """True si la celda está pintada de algo que el usuario pudo haber creído
    amarillo: un amarillo/dorado distinto del estándar, o un color del tema
    (que no se puede resolver sin leer el XML del tema del libro).

    Sólo se usa para avisar: la fila igual no se procesa.
    """
    fill = _fill_solido(cell)
    if fill is None:
        return False
    color = fill.fgColor
    if color.type == "theme":
        return True
    if color.type != "rgb":
        return False
    rgb = str(color.rgb or "").upper()[-6:]
    if len(rgb) != 6 or rgb == "FFFF00":
        return False
    try:
        r, g, b = (int(rgb[i:i + 2], 16) for i in (0, 2, 4))
    except ValueError:
        return False
    # Amarillos y dorados: rojo y verde altos, azul claramente por debajo.
    return r >= 180 and g >= 150 and r - b >= 40 and g - b >= 40


def _fila_con_datos(row, cols: dict[str, int]) -> bool:
    """Fila que tiene todo lo necesario para ser un pago (sin mirar el color)."""
    documento = str(_cel(row, cols.get("documento")) or "").strip()
    proveedor = str(_cel(row, cols.get("proveedor")) or "").strip()
    importe   = _importe(_cel(row, cols.get("importe")))
    return bool(documento and proveedor and importe is not None and importe != 0)


def leer_dm(path: Path | str, hoja: str = HOJA_DM,
            solo_amarillas: bool = True,
            avisos_out: list[str] | None = None) -> list[ProveedorTanda]:
    # data_only=True para leer valores calculados; read_only por defecto False (necesario para leer estilos de celda)
    wb = load_workbook(str(path), data_only=True)
    if hoja not in wb.sheetnames:
        hojas_disp = ", ".join(wb.sheetnames)
        wb.close()
        raise ValueError(
            f"No se encontró la hoja «{hoja}» en el archivo.\n"
            f"Hojas disponibles: {hojas_disp}"
        )
    ws = wb[hoja]

    rows = ws.iter_rows()
    header_row = next(rows)
    cols = _detectar_columnas(header_row, avisos_out)

    _COLS_REQUERIDAS = {
        "documento":  "Documento (ej. «FC - 21562»)",
        "proveedor":  "Proveedor",
        "importe":    "Importe",
        "pago":       "Forma de pago",
    }
    faltantes = [label for key, label in _COLS_REQUERIDAS.items() if key not in cols]
    if faltantes:
        detectados = [
            str(c.value).strip() for c in header_row if str(c.value or "").strip()
        ]
        wb.close()
        leidos = (
            f"Encabezados leídos en la fila 1: {', '.join(detectados[:15])}"
            + ("…" if len(detectados) > 15 else "")
            if detectados
            else "La primera fila de la hoja está vacía."
        )
        raise ValueError(
            f"Faltan columnas requeridas en la hoja «{hoja}»:\n"
            + "\n".join(f"  • {f}" for f in faltantes)
            + "\n\nLos encabezados tienen que estar en la PRIMERA fila de la hoja "
              "(no puede haber títulos ni filas en blanco arriba).\n"
            + leidos
        )

    proveedores: dict[str, ProveedorTanda] = {}
    filas_casi_amarillas: list[int] = []

    for row in rows:
        if not any(c.value for c in row):
            continue
        if solo_amarillas and not _es_fila_amarilla(row):
            # Fila completa pintada de un color parecido al amarillo: el usuario
            # probablemente quiso marcarla para pagar y no se va a procesar.
            if _fila_con_datos(row, cols) and any(_color_casi_amarillo(c) for c in row):
                filas_casi_amarillas.append(row[0].row)
            continue

        # Se normaliza acá una sola vez: de este texto dependen la aplicación del
        # pago en el POST, la verificación de saldo y la detección de facturas.
        documento = normalizar_doc(_cel(row, cols.get("documento")))
        if not documento:
            continue

        proveedor_nombre = str(_cel(row, cols.get("proveedor")) or "").strip()
        if not proveedor_nombre:
            continue

        importe_raw = _importe(_cel(row, cols.get("importe")))
        if importe_raw is None or importe_raw == 0:
            continue
        # Inversión universal del signo: el signo del Excel determina si adeuda
        # (negativo en Excel → positivo interno = a pagar) o si es crédito
        # (positivo en Excel → negativo interno = saldo a favor).
        # Esto aplica a TODO tipo de documento (FC, ND, NC, PAGO, MOVFONDOS).
        importe = -importe_raw

        cuit_raw = str(_cel(row, cols.get("cuit")) or "").strip()
        cuit = _limpiar_cuit(cuit_raw) if cuit_raw else ""

        item = ItemFactura(
            documento=documento,
            comprobante=str(_cel(row, cols.get("comprobante")) or "").strip(),
            descripcion="",
            importe=importe,
            fecha_vto=_fecha(_cel(row, cols.get("fecha_vto"))),
            modalidad_pago=str(_cel(row, cols.get("pago")) or "").strip(),
        )

        # A4: agrupar por CUIT cuando está disponible; fallback a nombre
        clave = cuit if cuit else proveedor_nombre
        if clave not in proveedores:
            proveedores[clave] = ProveedorTanda(
                cuit=cuit,
                nombre=proveedor_nombre,
            )
        else:
            existing = proveedores[clave]
            if cuit and not existing.cuit:
                existing.cuit = cuit
        proveedores[clave].items.append(item)

    wb.close()

    if filas_casi_amarillas and avisos_out is not None:
        muestra = ", ".join(str(n) for n in filas_casi_amarillas[:10])
        avisos_out.append(
            f"• {len(filas_casi_amarillas)} fila(s) con datos completos están pintadas "
            f"de un color que NO es el amarillo estándar (fila {muestra}"
            f"{'…' if len(filas_casi_amarillas) > 10 else ''}) y por eso NO se "
            f"procesaron. Si hay que pagarlas, repintalas con el amarillo de "
            f"«Colores estándar» y volvé a cargar el archivo."
        )

    return [clasificar(p) for p in proveedores.values()]
