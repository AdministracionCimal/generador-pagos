from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from PyQt6.QtCore import Qt
from PyQt6.QtGui import QBrush, QColor
from PyQt6.QtWidgets import (
    QDialog, QDialogButtonBox, QFileDialog, QFrame, QHBoxLayout, QHeaderView,
    QLabel, QPushButton, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget,
)

from src.ui import theme

_ESTADO_VARIANT = {"OK": "success", "ERROR": "danger", "MANUAL": "warning"}
_ROW_ERROR_BG   = QColor("#FEF2F2")


def _fmt(importe: float) -> str:
    return f"$ {importe:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


class ResultDialog(QDialog):
    def __init__(self, resultados: list[dict], parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Resultado del procesamiento")
        self.resize(860, 560)
        self._resultados = resultados

        ok_items  = [r for r in resultados if r["estado"] == "OK"]
        err_items = [r for r in resultados if r["estado"] == "ERROR"]

        ok  = len(ok_items)
        err = len(err_items)

        total_ok  = sum(float(r.get("importe", 0) or 0) for r in ok_items)
        total_err = sum(float(r.get("importe", 0) or 0) for r in err_items)

        # — Header ————————————————————————————————————————————————————————
        title    = QLabel("Resultado del procesamiento")
        title.setObjectName("PageTitle")
        subtitle = QLabel(f"Se procesaron {len(resultados)} órdenes.")
        subtitle.setObjectName("PageSubtitle")

        # — Stats bar (3 KPIs) ——————————————————————————————————————————
        stats_bar = self._build_stats_bar([
            ("PROCESADAS OK",    ok,   _fmt(total_ok),                  "success"),
            ("CON ERROR",        err,  _fmt(total_err) if err else "—", "danger"),
            ("TOTAL CONFIRMADO", None, _fmt(total_ok),                  "neutral"),
        ])

        # — Tabla ————————————————————————————————————————————————————————
        self._tabla = QTableWidget(len(resultados), 5)
        self._tabla.setHorizontalHeaderLabels(
            ["Proveedor", "Estado", "OP prevista", "OP real / detalle", "Importe"]
        )
        self._tabla.setAlternatingRowColors(True)
        self._tabla.setShowGrid(False)
        self._tabla.verticalHeader().setVisible(False)
        self._tabla.verticalHeader().setDefaultSectionSize(34)
        self._tabla.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._tabla.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        hh = self._tabla.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)   # ancho fijo para badge
        hh.resizeSection(1, 80)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)

        _LEFT  = Qt.AlignmentFlag.AlignLeft  | Qt.AlignmentFlag.AlignVCenter
        _RIGHT = Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter

        for row, r in enumerate(resultados):
            is_error = r["estado"] == "ERROR"
            row_bg   = _ROW_ERROR_BG if is_error else None

            def _mk(text: str, align=_LEFT, bg=row_bg) -> QTableWidgetItem:
                it = QTableWidgetItem(text)
                it.setTextAlignment(align)
                if bg is not None:
                    it.setBackground(QBrush(bg))
                return it

            self._tabla.setItem(row, 0, _mk(r.get("nombre", "")))
            variant = _ESTADO_VARIANT.get(r["estado"], "neutral")
            self._tabla.setCellWidget(row, 1, theme.make_badge(r["estado"].title(), variant))
            self._tabla.setItem(row, 2, _mk(r.get("numero_previsto", "") or "—"))
            # Para OK: numero_real = referencia de la OP (ej. «PAGO - 14062»)
            # Para ERROR/MANUAL: detalle = mensaje de error o motivo
            op_real = r.get("numero_real") or ""
            detalle = r.get("detalle", "") or ""
            self._tabla.setItem(row, 3, _mk(op_real if op_real else detalle))
            importe  = r.get("importe", "")
            imp_text = _fmt(float(importe)) if isinstance(importe, (int, float)) else str(importe)
            self._tabla.setItem(row, 4, _mk(imp_text, _RIGHT))

        # — Footer label ——————————————————————————————————————————————————
        parts = []
        if ok:  parts.append(f"✓  {ok} confirmada{'s' if ok != 1 else ''}")
        if err: parts.append(f"✗  {err} con error")
        footer_lbl = QLabel("  ·  ".join(parts) if parts else "")
        footer_lbl.setObjectName("Muted")

        # — Bottom bar ———————————————————————————————————————————————————
        btn_xlsx = QPushButton("Exportar Excel")
        btn_xlsx.clicked.connect(self._exportar_excel)
        btn_ok = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        btn_ok.button(QDialogButtonBox.StandardButton.Ok).setObjectName("Primary")
        btn_ok.button(QDialogButtonBox.StandardButton.Ok).setText("Cerrar")
        btn_ok.accepted.connect(self.accept)

        bar = QHBoxLayout()
        bar.setSpacing(10)
        bar.addWidget(btn_xlsx)
        bar.addStretch()
        bar.addWidget(btn_ok)

        root = QVBoxLayout(self)
        root.setContentsMargins(24, 22, 24, 18)
        root.setSpacing(12)
        head = QVBoxLayout()
        head.setSpacing(2)
        head.addWidget(title)
        head.addWidget(subtitle)
        root.addLayout(head)
        root.addWidget(stats_bar)
        root.addWidget(self._tabla, stretch=1)
        root.addWidget(footer_lbl)
        root.addLayout(bar)

    def _build_stats_bar(self, stats: list[tuple]) -> QFrame:
        bar = QFrame()
        bar.setObjectName("Card")
        row = QHBoxLayout(bar)
        row.setContentsMargins(0, 0, 0, 0)
        row.setSpacing(0)

        for i, (label, value, sub, variant) in enumerate(stats):
            cell = QWidget()
            v = QVBoxLayout(cell)
            v.setContentsMargins(24, 14, 24, 14)
            v.setSpacing(3)

            lbl = QLabel(label)
            lbl.setObjectName("KpiLabel")
            fg = theme._BADGE_VARIANTS[variant][0]

            v.addWidget(lbl)

            if value is not None:
                num = QLabel(str(value))
                num.setObjectName("KpiNumber")
                num.setStyleSheet(f"color: {fg};")
                v.addWidget(num)
                sub_lbl = QLabel(sub)
                sub_lbl.setStyleSheet(f"font-size: 11px; color: {theme.TEXT_MUTED};")
                v.addWidget(sub_lbl)
            else:
                # TOTAL CONFIRMADO: solo el monto grande
                num = QLabel(sub)
                num.setObjectName("KpiNumber")
                num.setStyleSheet(f"color: {theme.TEXT_PRIMARY};")
                v.addWidget(num)

            row.addWidget(cell, stretch=1)

            if i < len(stats) - 1:
                sep = QFrame()
                sep.setFrameShape(QFrame.Shape.VLine)
                sep.setStyleSheet(
                    f"color: {theme.BORDER}; background: {theme.BORDER};"
                    f"max-width: 1px; margin: 14px 0;"
                )
                row.addWidget(sep)

        return bar

    def _exportar_excel(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "Guardar reporte", "resultado_pagos.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Resultado"

        # ── Estilos ──────────────────────────────────────────────────────
        _HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
        _ERROR_FILL  = PatternFill("solid", fgColor="FEE2E2")
        _TOTAL_FILL  = PatternFill("solid", fgColor="EFF6FF")
        _HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
        _BOLD        = Font(bold=True, size=10)
        _NUM_FMT     = '#.##0,00'   # miles con punto, decimales con coma

        headers = ["Proveedor", "Estado", "OP Prevista", "OP Real / Referencia", "Importe"]
        col_widths = [40, 12, 22, 35, 18]

        # ── Encabezado ───────────────────────────────────────────────────
        for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = _HEADER_FONT
            cell.fill      = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 20

        # ── Filas de datos ───────────────────────────────────────────────
        total_ok = 0.0
        for data_row, r in enumerate(self._resultados, start=2):
            estado   = r.get("estado", "")
            op_real  = r.get("numero_real") or ""
            detalle  = r.get("detalle", "") or ""
            referencia = op_real if op_real else detalle
            importe  = r.get("importe", None)
            imp_val  = float(importe) if isinstance(importe, (int, float)) else None

            if estado == "OK" and imp_val is not None:
                total_ok += imp_val

            ws.cell(row=data_row, column=1, value=r.get("nombre", ""))
            ws.cell(row=data_row, column=2, value=estado)
            ws.cell(row=data_row, column=3, value=r.get("numero_previsto", "") or "")
            ws.cell(row=data_row, column=4, value=referencia)
            imp_cell = ws.cell(row=data_row, column=5, value=imp_val)
            imp_cell.number_format = _NUM_FMT
            imp_cell.alignment = Alignment(horizontal="right")

            if estado == "ERROR":
                for col in range(1, 6):
                    ws.cell(row=data_row, column=col).fill = _ERROR_FILL

        # ── Fila de totales ───────────────────────────────────────────────
        total_row = len(self._resultados) + 2
        ws.cell(row=total_row, column=1, value="TOTAL CONFIRMADO").font = _BOLD
        tot_cell = ws.cell(row=total_row, column=5, value=total_ok)
        tot_cell.font         = _BOLD
        tot_cell.number_format = _NUM_FMT
        tot_cell.alignment    = Alignment(horizontal="right")
        for col in range(1, 6):
            ws.cell(row=total_row, column=col).fill = _TOTAL_FILL

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(path)
